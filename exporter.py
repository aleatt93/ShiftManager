import csv
import datetime
import calendar
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

WEEKDAYS = ["Lun", "Mar", "Mer", "Gio", "Ven", "Sab", "Dom"]
MONTHS = [0, "Gennaio", "Febbraio", "Marzo", "Aprile", "Maggio", "Giugno", "Luglio", "Agosto",
          "Settembre", "Ottobre", "Novembre", "Dicembre"]


class Exporter:
    """Gestisce gli export della programmazione turni"""
    def __init__(self,
                 schedule_data,
                 employees_list,
                 year: int,
                 month: int,
                 config):

        self.schedule_data = schedule_data
        self.employee_list = employees_list
        self.year = year
        self.month = month
        self.config = config
        self.employee_lookup = {emp.id: emp for emp in self.employee_list}

        self.SHIFT_CORRISPONDANCE = self.config["shift_settings"]["shift_representation"]

        # Preparazione di un data grid leggibile
        self.data_grid = self._prepare_data_grid()

    def _prepare_data_grid(self):
        """Converts the raw schedule data into a simple list of lists (a grid) that can be easily written
        to any file format."""

        grid = []
        num_days = calendar.monthrange(self.year, self.month)[1]

        # Creazione Header Row
        headers = ["Matricola", "Cognome", "Nome"]
        for day in range(1, num_days + 1):
            day_name = WEEKDAYS[datetime.date(self.year, self.month, day).weekday()]
            headers.append(f"{day_name} {day}")
        grid.append(headers)

        # Creazione Data Rows
        for employee in self.employee_list:
            row_values = [employee.serial_number, employee.surname, employee.name]
            for day in range(1, num_days + 1):
                current_date = datetime.date(self.year, self.month, day)
                shift_text = ""

                if current_date in employee.days_off:
                    shift_text = self.SHIFT_CORRISPONDANCE["off_duty"]
                elif self.schedule_data and current_date in self.schedule_data:
                    daily_shifts = self.schedule_data[current_date]
                    # Compare by employee ID instead of object identity (temp list has deep copies)
                    employee_ids_in_mattina_rep = [emp.id for emp in daily_shifts["mattina_rep"]]
                    if employee.id in employee_ids_in_mattina_rep:
                            shift_text = self.SHIFT_CORRISPONDANCE["mattina_rep"]
                    else:
                        for shift_type, emp_assigned_to_shift in daily_shifts.items():
                            # Compare by ID instead of object identity
                            employee_ids_in_shift = [emp.id for emp in emp_assigned_to_shift]
                            if employee.id in employee_ids_in_shift:
                                shift_text = self.SHIFT_CORRISPONDANCE.get(shift_type, "?")
                                break

                row_values.append(shift_text)
            grid.append(row_values)
        return grid

    def export_to_txt(self, filepath):
        """Exports the data grid to a formatted TXT file."""
        if not self.data_grid:
            return None

        # Calculate columns width
        num_of_columns = len(self.data_grid[0])
        columns_width = [0] * num_of_columns

        for row in self.data_grid:
            for i, cell in enumerate(row):
                cell_width = max(len(line) for line in cell.split("\n"))
                if cell_width > columns_width[i]:
                    columns_width[i] = cell_width

        # +++ LINES BUILDING +++
        lines = []

        # Header Row
        header_cells = []
        for i, header in enumerate(self.data_grid[0]):
            # Ceneter header text within calculated column width
            header_cells.append(header.replace("\n", " ").center(columns_width[i] + 2))
        lines.append(" | ".join(header_cells))

        # Separator Line
        separator_cells = ["-" * (w + 2) for w in columns_width]
        lines.append("-+-".join(separator_cells))

        # Data Rows
        for row in self.data_grid[1:]:
            data_cells = []
            for i, cell in enumerate(row):
                # Center the data text
                data_cells.append(cell.center(columns_width[i] + 2))
            lines.append(" | ".join(data_cells))

        # Write content to the file
        file_content = "\n".join(lines) # Join all prepared lines with a newline character

        with open(filepath, "w", encoding="utf-8") as f:
            f.write(file_content)
            return None

    def export_to_csv(self, filepath):
        """Exports the data grid to a CSV file."""

        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            writer=csv.writer(f)

            # CSV doesn't handle multi-line headers well, so we replace newline with a space.
            header = [h.replace('\n', ' ') for h in self.data_grid[0]]
            writer.writerow(header)
            writer.writerows(self.data_grid[1:])

    def export_to_xlsx(self, filepath):
        """exports the data grid to an Excel (.xlsx) file."""

        number_of_rows = len(self.data_grid)
        wb = Workbook()
        ws = wb.active
        ws.title = f"Turni {MONTHS[self.month]} {self.year}"

        for row_data in self.data_grid:
            ws.append(row_data)

        # Styling
        header_font = Font(bold=True)
        center_alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )

        for cell in ws[1]:
            cell.font = header_font # Make headers bold

        for row in range(1, number_of_rows + 1):
            for cell in ws[row]:
                cell.alignment = center_alignment # Center alignment in all cells

            # Adjust column widths
            ws.column_dimensions["A"].width = 20
            ws.column_dimensions["B"].width = 20
            ws.column_dimensions["C"].width = 20

            wb.save(filepath)